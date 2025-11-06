import os
import re
import sys
from io import StringIO
from pathlib import Path
from dotenv import load_dotenv
import pandas as pd
import numpy as np
import google.generativeai as genai
from flask import Flask, render_template, request, jsonify, session
import openpyxl
from openpyxl import Workbook
import ast  # <<< para literal_eval seguro

# >>> Upload
from werkzeug.datastructures import FileStorage
import uuid

# =========================
# Estado global / Paths
# =========================
ULTIMO_TRACE = {}
BASE_DIR = Path(__file__).parent
DATA_PATH = (BASE_DIR.parent / "data").resolve()

app = Flask(__name__)
app.secret_key = os.urandom(24)

# Limite de upload (20MB)
app.config['MAX_CONTENT_LENGTH'] = 20 * 1024 * 1024

dataframe_completo: pd.DataFrame | None = None
MODELO_GEMINI: genai.GenerativeModel | None = None
STATUS_DADOS = "CARREGANDO..."

# =========================
# Config Gemini
# =========================
try:
    load_dotenv(dotenv_path=BASE_DIR / '.env')
    api_key = os.getenv("GOOGLE_API_KEY")
    if not api_key:
        raise ValueError("Chave da API do Google não encontrada no .env")
    genai.configure(api_key=api_key)
    MODELO_GEMINI = genai.GenerativeModel('gemini-2.5-flash')
    print("API e Modelo do Gemini configurados com sucesso.")
except Exception as e:
    print(f"Erro fatal ao configurar a API: {e}")
    sys.exit(1)

# =========================
# Filtro de linhas frágeis
# =========================
FRAGILE_PATTERNS = [
    r'\bpens(e|ar)\b', r'\bracioc[ií]nio\b', r'\bcadeia de pensamento\b',
    r'\bcorrente de pensamento\b', r'\bdeliber(a|e)\b', r'\bpasso a passo\b',
    r'\ban[aá]lis(e|ar)\b', r'\bethapa\s*\d+\b', r'chain[- ]?of[- ]?thought',
    r'inner monologue', r'\bCOT\b'
]

def _filter_fragile_lines(text: str) -> str:
    if not isinstance(text, str) or not text.strip():
        return text or ''
    out = []
    for ln in text.splitlines():
        low = ln.lower()
        if any(re.search(pat, low, flags=re.IGNORECASE) for pat in FRAGILE_PATTERNS):
            continue
        out.append(ln)
    return "\n".join(out).strip()

def _sanitize_trace(raw: dict) -> dict:
    def _from_block(block: str) -> str:
        if not isinstance(block, str) or not block.strip():
            return ""
        m = re.search(r"```python\s*(.*?)\s*```", block, re.DOTALL | re.IGNORECASE)
        return m.group(1).strip() if m else block.strip()

    if not isinstance(raw, dict):
        return {"codigo_gerado": "", "resultado_codigo": "", "resposta_final": ""}

    codigo = raw.get("codigo_gerado") or _from_block(raw.get("bloco_codigo", ""))

    return {
        "codigo_gerado": codigo or "",
        "resultado_codigo": raw.get("resultado_codigo", "") or "",
        "resposta_final": raw.get("resposta_final", "") or "",
    }

# =========================
# Convenções do Excel / Projetos
# =========================
COLS_PADRAO = [
    'Número', 'Classificação', 'Categoria', 'Fase', 'Condição',
    'Nome', 'Duração', 'Como Fazer', 'Documento Referência', '% Concluída'
]

DATA_PATH.mkdir(parents=True, exist_ok=True)

def _safe_xlsx_name(nome: str) -> str:
    base = re.sub(r'[^A-Za-z0-9_.-]+', '_', (nome or '').strip())
    if not base:
        raise ValueError("Informe um nome válido para o projeto.")
    if not base.lower().endswith('.xlsx'):
        base += '.xlsx'
    return base

def create_project(nome: str) -> str:
    """
    Cria um arquivo .xlsx em root/data com aba 'Backlog' e COLS_PADRAO.
    """
    fname = _safe_xlsx_name(nome)
    path = (DATA_PATH / fname)
    if path.exists():
        raise FileExistsError(f"Já existe um projeto com o nome '{nome}'.")
    wb = Workbook()
    ws = wb.active
    ws.title = 'Backlog'
    ws.append(COLS_PADRAO)
    wb.save(path)
    return str(path)

def list_projects() -> list:
    """Lista nomes (sem extensão) dos projetos em root/data."""
    return sorted(p.stem for p in DATA_PATH.glob("*.xlsx"))

def reload_data() -> int:
    """
    Recarrega o DataFrame global a partir de root/data.
    """
    global dataframe_completo
    try:
        df_new = carregar_dados_excel(DATA_PATH)
        dataframe_completo = df_new
        return 0 if (df_new is None or df_new.empty) else len(df_new)
    except Exception as e:
        print(f"Erro ao recarregar dados: {e}")
        return 0

# >>> helper para validar extensão
def _allowed_xlsx(filename: str) -> bool:
    return str(filename).lower().endswith(".xlsx")

# =========================
# Carregamento de dados (Excel + hyperlinks)
# =========================
@app.before_request
def load_data_once():
    global dataframe_completo, STATUS_DADOS

    if dataframe_completo is None:
        print("Iniciando carregamento de dados (primeira requisição)...")
        STATUS_DADOS = "ANALISANDO DADOS..."
        try:
            dataframe_completo = carregar_dados_excel(DATA_PATH)
            if dataframe_completo is not None and not dataframe_completo.empty:
                STATUS_DADOS = "ONLINE"
                print(f"Dados carregados. Total de {len(dataframe_completo)} linhas.")
            else:
                STATUS_DADOS = "ERRO: SEM DADOS"
                print("AVISO: Nenhum dado carregado. Verifique a pasta 'data'.")
        except Exception as e:
            STATUS_DADOS = "ERRO NO SERVIDOR"
            print(f"Erro crítico ao carregar dados: {e}")

def carregar_dados_excel(pasta_dados: Path) -> pd.DataFrame | None:
    """
    Lê todos os .xlsx de root/data, concatena, e extrai hyperlinks da coluna
    'Documento Referência' (aceitando variações com/sem acento e underscore).
    Escreve a URL extraída em 'url_referencia'.
    """
    caminho_pasta = pasta_dados if isinstance(pasta_dados, Path) else Path(pasta_dados)
    if not caminho_pasta.is_dir():
        print(f"Erro: Pasta '{pasta_dados}' não encontrada.")
        return None

    arquivos_excel = list(caminho_pasta.glob("*.xlsx"))
    if not arquivos_excel:
        print(f"Aviso: Nenhum arquivo .xlsx encontrado em '{pasta_dados}'.")
        return None

    lista_de_dataframes = []
    print("Carregando arquivos Excel e extraindo links...")

    for arquivo in arquivos_excel:
        try:
            # Pandas lê valores
            dicionario_de_abas_pd = pd.read_excel(arquivo, sheet_name=None, engine='openpyxl')
            # openpyxl lê metadados (hyperlinks)
            wb = openpyxl.load_workbook(arquivo, data_only=True)

            for nome_da_aba, df_aba in dicionario_de_abas_pd.items():
                if nome_da_aba not in wb.sheetnames:
                    continue

                ws = wb[nome_da_aba]

                # Mapeia cabeçalhos -> índice numérico da coluna
                header_map = {}
                for cell in ws[1]:
                    key = str(cell.value).strip() if cell.value is not None else ""
                    if key:
                        header_map[key] = cell.col_idx  # inteiro correto

                # Normalização para localizar "Documento Referência" em variações
                header_lower = {str(k).lower(): v for k, v in header_map.items()}
                possiveis = (
                    "documento referência",
                    "documento_referência",
                    "documento referencia",
                    "documento_referencia",
                )
                col_key = None
                for probe in possiveis:
                    if probe in header_lower:
                        col_key = header_lower[probe]
                        break

                links_map = {}
                if col_key is not None:
                    # i=0 corresponde à primeira linha de dados do DF (Excel linha 2)
                    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row,
                                                         min_col=col_key, max_col=col_key)):
                        cell = row[0]
                        if cell.hyperlink:
                            links_map[i] = cell.hyperlink.target

                # Anotações úteis
                df_aba['fonte_do_arquivo'] = arquivo.name
                df_aba['aba_do_projeto'] = nome_da_aba
                df_aba['url_referencia'] = df_aba.index.map(links_map)

                lista_de_dataframes.append(df_aba)

            wb.close()
            print(f"  - Arquivo '{arquivo.name}' lido e links processados.")

        except Exception as e:
            print(f"  - ERRO ao ler o arquivo {arquivo.name}: {e}")

    if not lista_de_dataframes:
        print("Nenhum dataframe foi carregado.")
        return None

    print("Concatenando todos os dataframes...")
    df_final = pd.concat(lista_de_dataframes, ignore_index=True)
    df_final['url_referencia'] = df_final['url_referencia'].fillna(value=np.nan)
    return df_final

# =========================
# Agente
# =========================
def executar_agente(df: pd.DataFrame, pergunta: str, historico: list) -> str:
    global MODELO_GEMINI, ULTIMO_TRACE
    if MODELO_GEMINI is None:
        return "Erro: O modelo de IA não foi inicializado."

    amostra_df = df.head().to_string()
    tipos_de_dados = df.dtypes.to_string()

    prompt_agente = f"""
        Você é um assistente de análise de dados Python Sênior.
        Você tem acesso a um DataFrame do Pandas chamado `df`.

        A estrutura do DataFrame (as primeiras 5 linhas) é a seguinte:
        {amostra_df}

        Os tipos de dados (dtypes) das colunas são:
        {tipos_de_dados}

        Histórico da conversa anterior (para contexto):
        {historico}

        Tarefa: Responda à pergunta do usuário: "{pergunta}"

        REGRAS DE CÓDIGO IMPORTANTES:
        1. A coluna 'Como Fazer' OU 'como_fazer' contém o TEXTO (ex: "Texto.1").
        2. A coluna 'Documento Referência' OU contém um LINK ou HYPERLINK (ex: "http://...").
        3. Se a pergunta pedir por um link ou documento, você DEVE usar as duas colunas.
        4. Seu código `print()` DEVE formatar a saída como: "TEXTO | URL"
        Exemplo de código:
        dados = df[df['alguma_coluna'] == 'valor_procurado'].iloc[0]
        print(f"{{dados['documento_referencia']}} | {{dados['url_referencia']}}")
        5. Contar tarefas usando as colunas 'condicao' e 'Condição' Elas podem ser traduzidas da seguinte maneira: 'Sempre' : 'Critica', 'A' : 'Alta', 'B' : 'Média', 'C' : 'Baixa'. 
        
        6. Uma tarefa está concluída se 'porcentagem' for 1.0 ou 100.0 OU se '% Concluída' for 1.0 ou 100.0.
        Usar .fillna(-1) para tratar NaNs de forma que não sejam contados como concluídos, ou em curso
        pois NaNs não são 1.0 nem 100.0.
        7. Uma tarefa está em curso ou em progresso se 'em curso' for 1.0 OU se 'status' for 'em curso'
        Usar .fillna(-1) para tratar NaNs de forma que não sejam contados como concluídos ou em curso, pois não envolve os valores procurados  

        8. Tarefas proximas do prazo são aquelas que tem o menor valor numérico na coluna 'duracao'
        
        9. Em caso de pesquisa ou procura de nomes, utilize as colunas 'nome' OU 'Nome' para encontrar a tarefa. Não utilize lower ou upper
            search_term = "nome da tarefa"
            # Realiza a pesquisa nas colunas 'nome' e 'Nome', utilizando .astype(str) para garantir comparação correta
            # e evitar erros com valores NaN.
            found_tasks = df[(df['nome'].astype(str) == search_term) | (df['Nome'].astype(str) == search_term)]
        10. Sempre que for perguntado sobre um nivel de condição/prioridade expecifico de uma tarefa  você deve analizar duas colunas: 'condicao' e 'Condição'
        11. Em caso de solicitação de colaboradores, seus emails estão armazenados na coluna 'colaboradores'
        12. Responsáveis pela tarefa estão em responsavel_conclusao OU em_curso_by
        13. Relatórios estão disponiveis em relatorio_progresso

        14. Sempre de algum tipo de resposta ou justificativa
        15. Priorize respostas bem formatadas
        
        Siga estritamente estas 3 etapas:

        ETAPA 1: Analise a pergunta, pense, e analise as regras acima.
        
        ETAPA 2: Escreva o código Python (dentro de ```python ... ```) para extrair os dados.
        - Use `print()` para o resultado

        ETAPA 3: Após o bloco de código, escreva a palavra "<STOP>".
        """

    # Helpers para sanitização e patch
    def _remove_redundant_imports(code: str) -> str:
        for pat in (
            r'^\s*import\s+pandas(\s+as\s+pd)?\s*$',
            r'^\s*from\s+pandas\b.*$',
            r'^\s*import\s+numpy(\s+as\s+np)?\s*$',
            r'^\s*from\s+numpy\b.*$',
        ):
            code = re.sub(pat, '', code, flags=re.MULTILINE)
        return code

    def _patch_nested_fstrings(code: str) -> str:
        """
        Correções heurísticas para casos frequentes de f-strings aninhadas.
        Ex.: f"... { f'{x} dias' }" -> "... " + (str(x) + " dias")
        e casos com closest_deadline_task['numeric_duracao'].
        """
        # substituições específicas
        code = code.replace(
            'f"{closest_deadline_task[\'numeric_duracao\']} dias"',
            'str(closest_deadline_task[\'numeric_duracao\']) + " dias"'
        )
        code = code.replace(
            "f'{closest_deadline_task[\"numeric_duracao\"]} dias'",
            "str(closest_deadline_task['numeric_duracao']) + ' dias'"
        )
        code = code.replace(
            "f\"{closest_deadline_task['numeric_duracao']} dias\"",
            "str(closest_deadline_task['numeric_duracao']) + ' dias'"
        )
        # padrão genérico simples: f'{expr} dias' -> str(expr) + ' dias'
        code = re.sub(
            r"f\"{([^{}]+)} dias\"",
            r"str(\1) + ' dias'",
            code
        )
        code = re.sub(
            r"f'\{([^{}]+)\} dias'",
            r"str(\1) + ' dias'",
            code
        )
        return code

    try:
        resposta_inicial = MODELO_GEMINI.generate_content(prompt_agente)
        bloco_codigo = re.search(r"```python\n(.*?)```", resposta_inicial.text, re.DOTALL)

        if not bloco_codigo:
            ULTIMO_TRACE = {
                "prompt_agente": prompt_agente,
                "bloco_codigo": None,
                "codigo_gerado": None,
                "resultado_codigo": None,
                "prompt_final": None,
                "resposta_final": None,
                "raw_primeira_resposta": resposta_inicial.text.strip(),
            }
            return resposta_inicial.text.replace("<STOP>", "").strip()

        codigo_gerado = bloco_codigo.group(1).strip()
        codigo_gerado = _remove_redundant_imports(codigo_gerado)

        print(f"--- Código Gerado (sanitizado) ---\n{codigo_gerado}\n---------------------")

        # Ambiente seguro para exec
        safe_builtins = {
            "print": print, "len": len, "range": range, "min": min, "max": max,
            "sum": sum, "abs": abs, "round": round, "sorted": sorted,
            "str": str, "int": int, "float": float, "list": list, "dict": dict, "set": set,
            "any": any, "all": all, "enumerate": enumerate, "zip": zip,
            "FileExistsError": FileExistsError, "ValueError": ValueError,
            "__import__": __import__,  # permite imports seguros no sandbox
        }

        safe_globals = {
            "__builtins__": safe_builtins,
            "df": df,
            "pd": pd,
            "np": np,
            "re": re,
            "ast": ast,
            "eval": ast.literal_eval,  # evita eval inseguro
            # Ferramentas de projeto expostas
            "create_project": create_project,
            "list_projects": list_projects,
            "reload_data": reload_data,
        }

        # Tenta compilar; se quebrar por f-string, aplica patch e recompila
        try:
            code_obj = compile(codigo_gerado, "<generated>", "exec")
        except SyntaxError as se:
            if "f-string" in str(se):
                print("Patch anti f-string aninhada aplicado...")
                codigo_gerado = _patch_nested_fstrings(codigo_gerado)
                code_obj = compile(codigo_gerado, "<generated_patched>", "exec")
            else:
                raise

        old_stdout = sys.stdout
        redirected_output = sys.stdout = StringIO()
        exec(code_obj, safe_globals, {})
        sys.stdout = old_stdout
        resultado_codigo = redirected_output.getvalue().strip()
        print(f"--- Resultado do Código ---\n{resultado_codigo}\n-------------------------")

        prompt_final = f"""
            Você é um assistente de dados amigável.
            A pergunta original do usuário foi: "{pergunta}"
            O código Python foi executado e o resultado (a saída do print) foi:
            {resultado_codigo}

            Com base *apenas* neste resultado, formule uma resposta clara e concisa em português.
            """
        resposta_final = MODELO_GEMINI.generate_content(prompt_final)
        resposta_txt = _filter_fragile_lines(resposta_final.text.strip())
        print(f"--- Resposta Final ---\n{resposta_txt}\n----------------------")

        ULTIMO_TRACE = {
            "prompt_agente": prompt_agente,
            "bloco_codigo": "```python\n" + codigo_gerado + "\n```",
            "codigo_gerado": codigo_gerado,
            "resultado_codigo": resultado_codigo,
            "prompt_final": prompt_final,
            "resposta_final": resposta_txt,
            "raw_primeira_resposta": resposta_inicial.text.strip(),
        }

        return resposta_txt

    except Exception as e:
        print(f"Erro no agente: {e}")
        ULTIMO_TRACE = {
            "prompt_agente": prompt_agente,
            "bloco_codigo": None,
            "codigo_gerado": None,
            "resultado_codigo": None,
            "prompt_final": None,
            "resposta_final": None,
            "raw_primeira_resposta": None,
            "erro": str(e),
        }
        return f"Desculpe, ocorreu um erro ao processar sua pergunta. Detalhes técnicos: {e}"

# =========================
# Rotas
# =========================
@app.route('/')
def index():
    session['messages'] = []
    session.modified = True
    return render_template(
        'index.html',
        messages=session['messages'],
        page_title="Chatbot Analista de Dados",
        status_text=STATUS_DADOS
    )

def _quer_debug(req) -> bool:
    if str(req.args.get("debug", "")).lower() in ("1", "true", "yes", "on"):
        return True
    if str(req.headers.get("X-Debug", "")).lower() in ("1", "true", "yes", "on"):
        return True
    if req.is_json:
        data = req.get_json(silent=True) or {}
        if str(data.get("debug", "")).lower() in ("1", "true", "yes", "on"):
            return True
    return False

@app.route('/chat', methods=['POST'])
def chat():
    global dataframe_completo, MODELO_GEMINI, ULTIMO_TRACE

    try:
        data = request.get_json(force=True, silent=False) or {}
        user_message = (data.get('message') or '').strip()
        if not user_message:
            return jsonify({"error": "Mensagem vazia."}), 400

        if 'messages' not in session:
            session['messages'] = []
        session['messages'].append({"role": "user", "content": user_message})

        if dataframe_completo is None or (hasattr(dataframe_completo, 'empty') and dataframe_completo.empty):
            raise ValueError("Os dados ainda não foram carregados. Aguarde.")

        historico_sessao = list(session['messages'])
        bot_response = executar_agente(dataframe_completo, user_message, historico_sessao)

        session['messages'].append({"role": "assistant", "content": bot_response})
        session.modified = True

        payload = {"reply": bot_response}
        # Enviando sempre o trace para o painel da esquerda:
        payload["trace"] = _sanitize_trace(ULTIMO_TRACE or {})

        return jsonify(payload)

    except Exception as e:
        print(f"Erro na rota /chat: {e}")
        return jsonify({"error": str(e)}), 500

# >>> APIs para listagem e upload de projetos (.xlsx)
@app.route('/api/projetos', methods=['GET'])
def api_listar_projetos():
    try:
        return jsonify({"projects": list_projects()})
    except Exception as e:
        app.logger.exception("Falha ao listar projetos")
        return jsonify({"error": str(e), "data_path": str(DATA_PATH)}), 500

@app.route('/api/upload-xlsx', methods=['POST'])
def api_upload_xlsx():
    """
    Recebe um XLSX e:
    - modo=substituir: grava por cima do projeto indicado em 'target'
    - modo=novo: cria um novo arquivo em root/data a partir do nome do anexo
    Sempre valida o arquivo e chama reload_data() ao final.
    """
    try:
        file: FileStorage | None = request.files.get('file')
        if file is None or not file.filename:
            return jsonify({"error": "Nenhum arquivo enviado."}), 400
        if not _allowed_xlsx(file.filename):
            return jsonify({"error": "Formato inválido. Envie um .xlsx"}), 400

        modo = (request.form.get('modo') or 'substituir').strip().lower()
        target = (request.form.get('target') or '').strip()

        # salva temporário na pasta data
        tmp_name = f"__upload_{uuid.uuid4().hex}.xlsx"
        tmp_path = DATA_PATH / tmp_name
        file.save(tmp_path)

        # valida abertura
        try:
            wb = openpyxl.load_workbook(tmp_path, data_only=True)
            wb.close()
        except Exception as e:
            tmp_path.unlink(missing_ok=True)
            return jsonify({"error": f"Arquivo inválido: {e}"}), 400

        action = None
        project_name = None

        if modo == 'substituir':
            if not target:
                tmp_path.unlink(missing_ok=True)
                return jsonify({"error": "Informe o projeto de destino em 'target'."}), 400
            dest = DATA_PATH / _safe_xlsx_name(target)
            # move (substitui)
            tmp_path.replace(dest)
            action = 'replace'
            project_name = dest.stem

        elif modo == 'novo':
            base_stem = Path(file.filename).stem
            dest = DATA_PATH / _safe_xlsx_name(base_stem)
            i = 1
            while dest.exists():
                dest = DATA_PATH / f"{base_stem}_{i}.xlsx"
                i += 1
            tmp_path.replace(dest)
            action = 'new'
            project_name = dest.stem

        else:
            tmp_path.unlink(missing_ok=True)
            return jsonify({"error": "Parâmetro 'modo' inválido. Use 'substituir' ou 'novo'."}), 400

        linhas = reload_data()
        return jsonify({"ok": True, "action": action, "project": project_name, "rows": linhas})

    except Exception as e:
        app.logger.exception("Falha no upload de XLSX")
        return jsonify({"error": str(e)}), 500

# =========================
# Main
# =========================
if __name__ == '__main__':
    print("Servidor Flask iniciando...")
    print("Acesse http://127.0.0.1:5000")
    print("Os dados serão carregados na primeira requisição ao site.")
    app.run(debug=True, port=5000)
