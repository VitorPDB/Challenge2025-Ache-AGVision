# app_final.py

# -----------------------------------------------------------------------------
# IMPORTS
# -----------------------------------------------------------------------------
from werkzeug.security import check_password_hash, generate_password_hash
from functools import wraps
import os
import re
import csv
import logging
import unicodedata
import difflib
from io import StringIO
import pandas as pd
from flask import Flask, request, jsonify, render_template, redirect, session, make_response, url_for
from flask_cors import CORS
from openpyxl import load_workbook
import re, logging, glob, datetime
import json
import numpy as np
import uuid
from seguranca_utils import (
    audit_log, rotate_backups, ensure_columns, assign_uuids,
    validate_row, find_row, set_runtime_refs, mutating, restore_task_from_audit, ValidationError
)
# --- Helpers para respostas JSON-safe (NaN -> None) -----------------------
import math

app = Flask(__name__, template_folder='templates', static_folder='static')
CORS(app)

from werkzeug.middleware.dispatcher import DispatcherMiddleware
from Chatbot.app import app as chatbot_app  # requer Chatbot/__init__.py

app.wsgi_app = DispatcherMiddleware(app.wsgi_app, {
    '/chatbot': chatbot_app,
    '/chat': chatbot_app,       # ← compat: o front chama /chat
    '/clear': chatbot_app,      # ← se o front chamar /clear
    '/status': chatbot_app      # ← se o front chamar /status
})

def _json_safe(value):
    """Converte NaN/NaT em None recursivamente em dict/list."""
    if isinstance(value, dict):
        return {k: _json_safe(v) for k, v in value.items()}
    if isinstance(value, list):
        return [_json_safe(v) for v in value]
    try:
        # pandas / float NaN
        if value is None:
            return None
        if isinstance(value, float) and math.isnan(value):
            return None
        # pandas NA/NaT
        import pandas as _pd
        if _pd.isna(value):
            return None
    except Exception:
        pass
    return value



# -----------------------------------------------------------------------------
# Configuração básica
# -----------------------------------------------------------------------------
# chave de sessão (use variável de ambiente em produção)
app.secret_key = os.environ.get('SECRET_KEY', 'dev-change-me')

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
USERS_FILE = os.path.join(BASE_DIR, 'data', 'users.json')
os.makedirs(os.path.dirname(USERS_FILE), exist_ok=True)




DATA_DIR = os.path.join(BASE_DIR, "data")        # <- NOVO
PDF_DIR  = os.path.join(BASE_DIR, "static", "pdfs")  # <- NOVO

os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(PDF_DIR, exist_ok=True)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)
# -----------------------------------------------------------------------------
# UTIL: carregar base Excel (tarefas)
# -----------------------------------------------------------------------------
# --------- Helpers de normalização ---------
_COLMAP = {
    'Número':'numero', 'Nº':'numero',
    'Fase':'fase',
    'Nome':'nome',
    'Categoria':'categoria',
    'Duração':'duracao',
    'Condição':'condicao', 'Prioridade':'condicao',
    'Concluida':'concluida', 'Concluída':'concluida',
    'Classificação':'classificacao',
    'Como Fazer':'como_fazer',
    'Documento Referência':'documento_referencia',
    '% Concluída':'porcentagem', '% Concluido':'porcentagem'
}

# UUID de projeto determinístico a partir do nome do arquivo (sem depender de disco)
def _project_uuid(nome_proj: str) -> str:
    return str(uuid.uuid5(uuid.NAMESPACE_DNS, f"challenge2025::{nome_proj}"))




def _slug_int(v, default=9999):
    #Extrai inteiro de strings tipo '5 dias', '10d', '15' etc.
    if pd.isna(v): return default
    if isinstance(v, (int, float)) and pd.notna(v): 
        try: return int(v)
        except: return default
    s = str(v).lower()
    if 'conclu' in s: return 'Concluído'
    m = re.search(r'(-?\d+)', s)
    return int(m.group(1)) if m else default

def _norm_percent(x):
    #Coerção %: aceita 0–100 ou 0–1 (se <=1 e tiver decimal, multiplica por 100).
    try:
        v = float(str(x).replace('%','').strip())
        if v <= 1 and ('.' in str(x) or ',' in str(x)):
            v *= 100
        return int(round(v))
    except:
        return 0

def _rename_cols(df):
    rename = {c: _COLMAP[c] for c in df.columns if c in _COLMAP}
    df = df.rename(columns=rename)
    return df

def _garantir_cols(df):
    base = [
        'numero','fase','nome','categoria','duracao','condicao',
        'concluida','porcentagem','Sheet','Projeto',
        'classificacao','como_fazer','documento_referencia',
        # --- ETAPA 2 ---
        'em_curso',              # 0/1
        'em_curso_by',           # e-mail do dono atual
        'inicio_em',             # timestamp início
        'colaboradores',         # csv ou json
        'relatorio_progresso',   # texto livre
        'responsavel_conclusao', # já usado no concluir
        'data_conclusao',        # idem
        'status'                 # Ex.: Em curso / Concluída
    ]
    for c in base:
        if c not in df.columns:
            if c in ('numero','duracao','concluida','porcentagem','em_curso'):
                df[c] = 0
            else:
                df[c] = ''
    return df

# --------- Carregador principal ---------
def carregar_base_dados(projeto=None):
    """
    Se projeto=None → concatena todos os projetos (todos .xlsx).
    Se projeto="Nome do arquivo sem .xlsx" → carrega só aquele.
    Sempre adiciona colunas 'Sheet' e 'Projeto'.

    Novidade:
      - texto_auxiliar: texto visível na célula "Documento Referência"
      - documento_auxiliar: URL do hyperlink embutido nessa célula (target)
        * Se não houver hyperlink, tenta usar o valor da célula se parecer uma URL
    """
    def _norm(s: str) -> str:
        import unicodedata, re
        s = str(s or '').strip()
        s = ''.join(ch for ch in unicodedata.normalize('NFKD', s) if not unicodedata.combining(ch))
        s = re.sub(r'\s+', ' ', s)
        return s.lower()

    # nomes possíveis (sem acento) para localizar a coluna no cabeçalho do Excel
    docref_headers_norm = {
        _norm('Documento Referência'),
        _norm('Documento Referencia'),
        _norm('Documento'),
        _norm('Doc Referência'),
        _norm('Doc Referencia'),
    }

    projetos = _listar_projetos()
    arquivos = [projetos[projeto]] if (projeto and projeto in projetos) else list(projetos.values())

    all_dfs = []
    for path in arquivos:
        nome_proj = os.path.splitext(os.path.basename(path))[0]
        try:
            # 1) Lê os dados em DataFrames (pandas)
            sheets = pd.read_excel(path, sheet_name=None, engine="openpyxl")

            # 2) Abre o mesmo arquivo com openpyxl para capturar hyperlinks
            from openpyxl import load_workbook
            wb = load_workbook(path, data_only=True)

            for aba, df in (sheets or {}).items():
                if df is None or df.empty:
                    continue

                # Captura colunas de hyperlink desta ABA
                texto_aux_list = []
                link_aux_list  = []

                try:
                    ws = wb[aba]
                    # Detecta linha de cabeçalho (assumindo a 1ª linha)
                    header_cells = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
                    # mapa: índice (0-based) → valor normalizado
                    header_norm_map = {i: _norm(v) for i, v in enumerate(header_cells)}
                    # encontra a(s) coluna(s) candidata(s) a "Documento Referência"
                    cand_cols = [i for i, nm in header_norm_map.items() if nm in docref_headers_norm]

                    # pega a primeira candidata, se houver
                    doc_col_idx = cand_cols[0] if cand_cols else None

                    # percorre as linhas de dados (a partir da 2ª linha)
                    if doc_col_idx is not None:
                        for r in ws.iter_rows(min_row=2, max_row=ws.max_row):
                            cell = r[doc_col_idx]
                            valor_visivel = cell.value if cell.value is not None else ''
                            link = ''
                            if cell.hyperlink and getattr(cell.hyperlink, 'target', ''):
                                link = cell.hyperlink.target
                            else:
                                # fallback: se o texto já for uma URL "crua", usa como link
                                vs = str(valor_visivel).strip()
                                import re as _re
                                if _re.match(r'^(https?://|www\.)', vs, flags=_re.I):
                                    link = vs

                            texto_aux_list.append(str(valor_visivel) if valor_visivel is not None else '')
                            link_aux_list.append(str(link))
                    else:
                        # não achou a coluna; mantém listas vazias (depois alinhamos pelo tamanho do DF)
                        pass
                except Exception:
                    # se der algo errado ao ler hyperlinks, segue sem travar
                    texto_aux_list, link_aux_list = [], []

                # 3) renomeia colunas e marca origem
                df = _rename_cols(df)
                df['Sheet']   = aba
                df['Projeto'] = nome_proj

                # 4) garante TODAS as colunas antes de normalizar
                df = _garantir_cols(df)
                # evita NaN nessas colunas usadas pela lógica de execução
                df['em_curso']     = pd.to_numeric(df['em_curso'], errors='coerce').fillna(0).astype(int)
                df['em_curso_by']  = df['em_curso_by'].fillna('').astype(str)
                df['concluida']    = pd.to_numeric(df['concluida'], errors='coerce').fillna(0).astype(int)

                df = ensure_columns(df, project_uuid=_project_uuid(nome_proj))
                df = assign_uuids(df)

                # 5) anexa colunas novas (alinha pelo comprimento do DF)
                n = len(df)
                if len(texto_aux_list) < n: texto_aux_list += [''] * (n - len(texto_aux_list))
                if len(link_aux_list)  < n: link_aux_list  += [''] * (n - len(link_aux_list))
                df['texto_auxiliar']     = texto_aux_list[:n]
                df['documento_auxiliar'] = link_aux_list[:n]

                # 6) normalizações antigas
                df['duracao']     = df['duracao'].apply(_slug_int)
                df['concluida']   = df['concluida'].fillna(0)
                df['concluida']   = df['concluida'].apply(lambda x: 1 if str(x).strip().lower() in
                                        ['1','100','sim','true','concluído','concluido'] else 0)
                df['porcentagem'] = df['porcentagem'].apply(_norm_percent)

                all_dfs.append(df)

        except Exception as e:
            logger.exception(f"Falha lendo {path}: {e}")

    if not all_dfs:
        # Retorna DF vazio com colunas padrão + novas
        return pd.DataFrame(columns=[
            'numero','fase','nome','categoria','duracao','condicao',
            'concluida','porcentagem','Sheet','Projeto',
            'classificacao','como_fazer','documento_referencia',
            'texto_auxiliar','documento_auxiliar'
        ])

    df = pd.concat(all_dfs, ignore_index=True)

    # coerção final: se 'duracao' tem "conclu", marca concluída/100%
    mask_conc = df['duracao'].astype(str).str.contains('conclu', case=False, na=False)
    df.loc[mask_conc, ['concluida','porcentagem']] = [1, 100]

    # garante as novas colunas caso alguma aba não tenha produzido
    for c in ['texto_auxiliar','documento_auxiliar']:
        if c not in df.columns: df[c] = ''

    return df

# -----------------------------------------------------------------------------
# PÁGINAS PRINCIPAIS
# -----------------------------------------------------------------------------
@app.route('/')
def homepage():
    return render_template('index.html')

@app.route('/supervisor')
def supervisor_page():
    return render_template('supervisor.html')

@app.route('/documentos')
def documentos():
    args = request.args.to_dict()
    return render_template('documentos.html', **args)

@app.route('/textos')
def textos():
    args = request.args.to_dict()
    return render_template('textos.html', **args)

@app.route('/controle')
def controle_page():
    return render_template('controle.html')

# -----------------------------------------------------------------------------
# CRONOGRAMA
# -----------------------------------------------------------------------------
import re, unicodedata
from flask import request, jsonify

def _fase_key(s: str) -> str:
    s = str(s or '').lower()
    s = ''.join(ch for ch in unicodedata.normalize('NFKD', s) if not unicodedata.combining(ch))
    s = re.sub(r'[^a-z0-9]+', ' ', s).strip()
    m = re.match(r'^(\d+)', s)
    return m.group(1) if m else s

@app.route('/gerar-cronograma', methods=['POST'])
def gerar_cronograma():
    try:
        data = request.get_json(force=True, silent=True) or {}
        projeto   = data.get('projeto')            # "" = todos os arquivos
        categoria = data.get('categoria') or ''    # opcional
        condicoes = data.get('condicoes') or {}    # dict { "1. Escopo...": ["Sempre","A"...], ... }

        # 1) carrega a base (um arquivo/projeto ou todos)
        df = carregar_base_dados(projeto)
        if df is None or df.empty:
            return jsonify({"cronograma": [], "duracao_total": 0})

        # 2) filtra por categoria se vier
        if categoria:
            df = df[df['categoria'].astype(str) == str(categoria)]

        # 3) filtra por condição por fase (case/acentos robustos)
        if isinstance(condicoes, dict) and condicoes:
            cond_map   = { _fase_key(k): set(v or []) for k, v in condicoes.items() }
            cond_union = set().union(*cond_map.values()) if cond_map else set()

            def tarefa_valida(row):
                fk   = _fase_key(row.get('fase', ''))
                cond = str(row.get('condicao', '')).strip()
                allowed = cond_map.get(fk, cond_union)
                return (cond in allowed) if allowed else True

            df = df[df.apply(tarefa_valida, axis=1)]

        # 4) garantir colunas mínimas
        ensure_str = ['como_fazer','documento_referencia','classificacao','Sheet','Projeto',
                      'fase','nome','categoria','condicao','texto_auxiliar','documento_auxiliar']
        ensure_num = ['numero','duracao','concluida','porcentagem']

        for col in ensure_str:
            if col not in df.columns: df[col] = ''
        for col in ensure_num:
            if col not in df.columns: df[col] = 0

        # 5) soma de duração
        df['duracao_num'] = pd.to_numeric(df['duracao'], errors='coerce').fillna(0).astype(int)
        duracao_total = int(df['duracao_num'].sum())

        # 6) ordem das colunas expostas no JSON
        ordem = [
            'numero','fase','nome','categoria','condicao','duracao','concluida',
            'porcentagem','Sheet','Projeto',
            # campos legacy
            'como_fazer','documento_referencia','classificacao',
            # novos campos para textos/documentos auxiliares
            'texto_auxiliar','documento_auxiliar', 'em_curso','em_curso_by','inicio_em','colaboradores','relatorio_progresso', 'responsavel_conclusao','data_conclusao','status'
        ]

        # garanta tipos razoáveis para evitar NaN → JSON inválido
        for col in ['numero','concluida','porcentagem','duracao']:
            df[col] = pd.to_numeric(df[col], errors='coerce')

        # 7) serialização estável
        cronograma_json = df[ordem] \
            .sort_values(by=['Projeto','fase','numero'], na_position='last') \
            .to_json(orient='records', force_ascii=False)
        cronograma = json.loads(cronograma_json)

        return jsonify({"cronograma": cronograma, "duracao_total": duracao_total})

    except Exception as e:
        logger.exception(f"/gerar-cronograma: {e}")
        return jsonify({"error": str(e), "cronograma": [], "duracao_total": 0}), 500

@app.get('/tarefas-em-curso')
def tarefas_em_curso():
    """
    Lista tarefas em_curso=1 (opcionalmente filtradas por projeto).
    GET /tarefas-em-curso?projeto=XYZ
    """
    projeto = request.args.get('projeto') or request.headers.get('X-Projeto')
    df = carregar_base_dados(projeto)
    if df is None or df.empty:
        return jsonify([])

    cols = [
        'Projeto','Sheet','numero','nome','fase','categoria','condicao',
        'em_curso','em_curso_by','inicio_em','porcentagem',
        'colaboradores','relatorio_progresso'
    ]
    for c in cols:
        if c not in df.columns:
            df[c] = ''

    # >>> robustez contra '', NaN, 'nan', etc.
    import pandas as pd
    df['em_curso'] = pd.to_numeric(df.get('em_curso', 0), errors='coerce').fillna(0).astype(int)

    emc = (
        df[df['em_curso'] == 1][cols]
        .sort_values(['Projeto','fase','numero'])
    )
    return jsonify(json.loads(emc.to_json(orient='records', force_ascii=False)))

# -----------------------------------------------------------------------------
# CONCLUIR / REABRIR TAREFA / INICIAR
# -----------------------------------------------------------------------------
@app.route('/concluir-tarefa', methods=['POST'])
def concluir_tarefa():
    """
    Concluir ou reabrir tarefa.

    Entrada (JSON):
      - numero: int | str
      - sheet: str
      - projeto: str (opcional; se ausente, tentamos descobrir)
      - acao: 'concluir' | 'reabrir'  (default: 'concluir')
      - version: int/str (opcional; lock otimista se existir coluna 'version')
      - operator: str (opcional; ou header 'X-Operator')
      - nova_duracao / duracao: int (opcional; prazo em dias para reabrir)
      - clear_owner: bool (opcional; ao reabrir, limpar "em_curso_by")  <<< NOVO

    Saída (JSON):
      - success: bool
      - error/details em caso de falha
      - task: dict (registro após alteração)
    """
    import pandas as pd
    from datetime import datetime

    try:
        data = request.get_json(force=True, silent=True) or {}

        # --------- parâmetros ---------
        numero   = data.get('numero')
        sheet    = data.get('sheet')
        projeto  = data.get('projeto') or request.headers.get('X-Projeto')
        acao     = (data.get('acao') or 'concluir').strip().lower()
        version_client = data.get('version')
        operator = data.get('operator') or request.headers.get('X-Operator') or 'system'

        # >>> NOVO: captura flag clear_owner
        clear_owner = bool(data.get('clear_owner') or data.get('limpar_responsavel') or False)

        # prazo para reabrir (aceita nova_duracao/duracao)
        novo_prz_raw = data.get('nova_duracao', data.get('duracao', 30))
        try:
            novo_prz = int(novo_prz_raw)
        except Exception:
            novo_prz = 30

        if numero is None or sheet is None:
            return jsonify({'success': False, 'error': 'Parâmetros insuficientes: numero e sheet são obrigatórios'}), 400

        # --------- descobrir projeto se não veio ---------
        def _descobrir_projeto_por_tarefa(num, aba):
            try:
                projetos = _listar_projetos()
            except Exception:
                return None
            hits = []
            for nome_proj, path in projetos.items():
                try:
                    dfp = carregar_base_dados(nome_proj)
                    if dfp is None or dfp.empty:
                        continue
                    m_num   = dfp.get('numero', pd.Series([], dtype=str)).astype(str) == str(num)
                    m_sheet = dfp.get('Sheet',  pd.Series([], dtype=str)).astype(str) == str(aba)
                    if (m_num & m_sheet).any():
                        hits.append(nome_proj)
                except Exception:
                    continue
            if not hits:
                return None
            if len(hits) > 1:
                return "__AMBIGUO__"
            return hits[0]

        if not projeto:
            projeto = _descobrir_projeto_por_tarefa(numero, sheet)
            if projeto is None:
                return jsonify({'success': False, 'error': 'Não encontrei esta tarefa em nenhum projeto.'}), 404
            if projeto == "__AMBIGUO__":
                return jsonify({'success': False, 'error': 'Tarefa encontrada em mais de um projeto; informe o campo "projeto".'}), 409

        # --------- carregar base do projeto ---------
        try:
            df = carregar_base_dados(projeto)
        except Exception as e:
            app.logger.error(f"Falha ao carregar base do projeto '{projeto}': {e}", exc_info=True)
            return jsonify({'success': False, 'error': f'Erro ao carregar dados do projeto {projeto}'}), 500

        if df is None or df.empty:
            return jsonify({'success': False, 'error': 'Base de tarefas vazia'}), 404

        # --------- localizar tarefa (robusto a tipos) ---------
        try:
            mask_num   = df['numero'].astype(str) == str(numero)
            mask_sheet = df['Sheet'].astype(str)  == str(sheet)
        except KeyError as e:
            return jsonify({'success': False, 'error': f'Coluna ausente na base: {e}'}), 500

        idxs = df.index[mask_num & mask_sheet].tolist()
        if not idxs:
            return jsonify({'success': False, 'error': 'Tarefa não encontrada'}), 404

        i = idxs[0]
        before = df.loc[i].to_dict()

        # --------- lock otimista (se coluna existir e cliente enviar) ---------
        if version_client is not None and 'version' in df.columns:
            try:
                if int(str(df.at[i, 'version'])) != int(str(version_client)):
                    return jsonify({'success': False, 'error': 'version_conflict', 'current': before}), 409
            except Exception:
                pass

        # --------- ajustes de dtype (duracao pode receber string "Concluído") ---------
        if 'duracao' in df.columns and df['duracao'].dtype != object:
            df['duracao'] = df['duracao'].astype('object')

        # --------- aplicar ação ---------
        agora = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        if acao == 'concluir':
            # marcador de conclusão
            if 'duracao' in df.columns:
                df.at[i, 'duracao'] = 'Concluído'
            if 'concluida' in df.columns:
                df.at[i, 'concluida'] = 1
            if 'em_curso' in df.columns:
                df.at[i, 'em_curso'] = 0
            if 'porcentagem' in df.columns:
                df.at[i, 'porcentagem'] = 100
            for col, val in [
                ('status', 'Concluída'),
                ('Status', 'Concluída'),
                ('data_conclusao', agora),
                ('responsavel_conclusao', operator),
            ]:
                if col in df.columns:
                    df.at[i, col] = val

            # garantir 100%
            if 'porcentagem' in df.columns:
                try:
                    df.at[i, 'porcentagem'] = 100
                except Exception:
                    df['porcentagem'] = df['porcentagem'].astype('object')
                    df.at[i, 'porcentagem'] = 100
            # info auxiliares (se existirem)
            for col, val in [
                ('status', 'Concluída'),
                ('Status', 'Concluída'),
                ('data_conclusao', agora),
                ('responsavel_conclusao', operator),
            ]:
                if col in df.columns:
                    df.at[i, col] = val

        elif acao == 'reabrir':
            # volta a ter prazo numérico e flags
            if 'duracao' in df.columns:
                df.at[i, 'duracao'] = novo_prz
            if 'concluida' in df.columns:
                df.at[i, 'concluida'] = 0
            if 'em_curso' in df.columns:
                df.at[i, 'em_curso'] = 1
            
            # >>> NOVO: limpar dono se solicitado
            if clear_owner and 'em_curso_by' in df.columns:
                df.at[i, 'em_curso_by'] = ''
            
            # tirar de 100% para reaparecer no Gestor
            if 'porcentagem' in df.columns:
                try:
                    cur = str(df.at[i, 'porcentagem']).strip().replace('%', '')
                    cur = int(cur) if cur else 0
                except Exception:
                    cur = 0
                df.at[i, 'porcentagem'] = 0 if cur >= 100 else cur
            for col, val in [
                ('status', 'Em curso'),
                ('Status', 'Em curso'),
                ('data_conclusao', ''),
                ('responsavel_conclusao', ''),
            ]:
                if col in df.columns:
                    df.at[i, col] = val

        else:
            return jsonify({'success': False, 'error': f'Ação inválida: {acao}'}), 400

        # --------- normalização/validação ---------
        after = df.loc[i].to_dict()
        try:
            try:
                after = normalize_row(after)
            except NameError:
                pass
            validate_row(after)
        except Exception as ve:
            details = getattr(ve, 'errors', None) or getattr(ve, 'args', [None])[0] or []
            # desfaz alterações na linha
            df.loc[i] = before
            return jsonify({'success': False, 'error': 'validation', 'details': details}), 400

        # --------- version++ ---------
        if 'version' in df.columns:
            try:
                curv = int(df.at[i, 'version']) if pd.notna(df.at[i, 'version']) else 0
            except Exception:
                curv = 0
            df.at[i, 'version'] = curv + 1

        after = df.loc[i].to_dict()
        safe_after = _json_safe(after)

        # --------- persistir no Excel do projeto ---------
        try:
            projetos = _listar_projetos()
            if projeto not in projetos:
                return jsonify({'success': False, 'error': 'Projeto inválido'}), 400
            target = projetos[projeto]

            try:
                rotate_backups(target)
            except Exception:
                pass

            with pd.ExcelWriter(target, engine='openpyxl') as w:
                # grava por aba (Sheet) apenas as linhas deste projeto
                df_proj = df[df['Projeto'].astype(str) == str(projeto)]
                for aba, gdf in df_proj.groupby('Sheet'):
                    cols = [c for c in gdf.columns if c not in ['Sheet', 'Projeto']]
                    gdf[cols].to_excel(w, sheet_name=str(aba), index=False)
        except Exception as e:
            app.logger.error(f"Erro ao salvar base: {e}", exc_info=True)
            return jsonify({'success': False, 'error': 'Erro ao salvar alterações'}), 500

        # --------- auditoria (não bloqueante) ---------
        try:
            audit_log({
                'action': acao,
                'operator': operator,
                'sheet': sheet,
                'task_uuid': after.get('task_uuid') or before.get('task_uuid'),
                'before': before,
                'after': safe_after,
                'ts': agora
            })
        except Exception:
            pass

        return jsonify({'success': True, 'projeto': projeto, 'task': after})

    except Exception as e:
        app.logger.error(f"Erro em concluir_tarefa: {e}", exc_info=True)
        return jsonify({'success': False, 'error': str(e)}), 500

@app.post('/iniciar-tarefa')
def iniciar_tarefa():
    """
    Marca uma tarefa como 'em_curso' por UM funcionário (exclusivo).
    Body JSON: { projeto?, sheet, numero, colaboradores?, relatorio?, porcentagem? }
    Regra:
      - se já estiver em curso por OUTRA pessoa -> 409
      - se estiver concluída -> 409 (ou reabrir via /concluir-tarefa com acao='reabrir')
    """
    from datetime import datetime
    data = request.get_json(force=True, silent=True) or {}
    numero   = data.get('numero')
    sheet    = data.get('sheet')
    projeto  = data.get('projeto') or request.headers.get('X-Projeto')
    colabs   = data.get('colaboradores')  # string CSV ou lista
    relato   = data.get('relatorio')
    pct_raw  = data.get('porcentagem')

    if numero is None or sheet is None:
        return jsonify({'success': False, 'error': 'Parâmetros insuficientes (numero, sheet)'}), 400

    # se projeto não vier, tentamos descobrir
    if not projeto:
        projeto = _descobrir_projeto_por_tarefa(int(numero), str(sheet))
        if projeto in (None, "__AMBIGUO__"):
            return jsonify({'success': False, 'error': 'Projeto não encontrado ou ambíguo; envie "projeto".'}), 409

    # carrega df do projeto
    df = carregar_base_dados(projeto)
    if df is None or df.empty:
        return jsonify({'success': False, 'error': 'Projeto sem dados'}), 404

    m = (df['numero'].astype(str) == str(numero)) & (df['Sheet'].astype(str) == str(sheet))
    idxs = df.index[m].tolist()
    if not idxs:
        return jsonify({'success': False, 'error': 'Tarefa não encontrada'}), 404
    i = idxs[0]

    # não permitir "Em curso" se já concluída
    try:
        is_conc = False
        dur_s   = str(df.at[i, 'duracao'])
        pct     = int(str(df.at[i, 'porcentagem']).replace('%', '') or 0)
        c_raw = pd.to_numeric([df.at[i, 'concluida']], errors='coerce')[0]
        conc  = int(c_raw) if pd.notna(c_raw) else 0
        if 'conclu' in dur_s.lower() or pct >= 100 or conc >= 1:
            is_conc = True
        if is_conc:
            return jsonify({'success': False, 'error': 'already_done', 'message': 'Tarefa já concluída'}), 409
    except Exception:
        pass

    op = _current_operator()

    dono_atual = str(df.at[i, 'em_curso_by'] or '').strip()
    ec_raw     = pd.to_numeric([df.at[i, 'em_curso']], errors='coerce')[0]
    em_curso   = int(ec_raw) if pd.notna(ec_raw) else 0
    if not op or op in ('system', ''):
        app.logger.warning(f"❌ Tentativa de iniciar tarefa sem operador válido. Header: {request.headers.get('X-Operator')}, Session: {session.get('user')}")
        return jsonify({
            'success': False,
            'error': 'no_operator',
            'message': 'Operador não identificado. Faça login e recarregue a página.'
        }), 401
    # exclusividade: se já estiver em curso por outra pessoa => 409
    if em_curso == 1 and dono_atual and dono_atual.lower() != op.lower():
        return jsonify({'success': False, 'error': 'locked_by_other',
                        'em_curso_by': dono_atual}), 409

    # marca como em curso pelo operador atual
    df.at[i, 'em_curso']    = 1
    df.at[i, 'em_curso_by'] = op
    if not str(df.at[i, 'inicio_em']).strip():
        df.at[i, 'inicio_em'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    df.at[i, 'status'] = 'Em curso'

    # colaboradores e relatório (optativos)
    if isinstance(colabs, list):
        colabs = ','.join([str(x).strip() for x in colabs if str(x).strip()])
    if colabs is not None:
        df.at[i, 'colaboradores'] = str(colabs).strip()
    if relato is not None:
        df.at[i, 'relatorio_progresso'] = str(relato).strip()

    # porcentagem (tolerante)
    if pct_raw is not None:
        try:
            p = int(str(pct_raw).replace('%','').strip() or 0)
        except Exception:
            p = 0
        df.at[i, 'porcentagem'] = max(0, min(99, p))  # evita "concluir" por 100

    # persistência por projeto/aba
    try:
        projetos = _listar_projetos()
        target   = projetos[projeto]
        with pd.ExcelWriter(target, engine='openpyxl') as w:
            df_proj = df[df['Projeto'].astype(str) == str(projeto)]
            for aba, gdf in df_proj.groupby('Sheet'):
                cols = [c for c in gdf.columns if c not in ['Sheet','Projeto']]
                gdf[cols].to_excel(w, sheet_name=str(aba), index=False)
    except Exception as e:
        app.logger.error(f"Erro ao salvar INICIAR: {e}", exc_info=True)
        return jsonify({'success': False, 'error': 'Erro ao salvar'}), 500

    try:
        audit_log({
            'action': 'iniciar',
            'operator': op,
            'sheet': sheet,
            'task_uuid': df.at[i, 'task_uuid'] if 'task_uuid' in df.columns else None,
            'before': None,
            'after': df.loc[i].to_dict(),
            'ts': df.at[i, 'inicio_em']
        })
    except Exception:
        pass

    return jsonify({'success': True, 'task': _json_safe(df.loc[i].to_dict()), 'projeto': projeto})


@app.post('/atualizar-progresso')
def atualizar_progresso():
    """
    Atualiza % de conclusão, colaboradores e relatório de uma tarefa que está em curso.
    Restringe edição ao 'dono' atual (em_curso_by) — salvo se vazio.
    """
    data = request.get_json(force=True, silent=True) or {}
    numero  = data.get('numero')
    sheet   = data.get('sheet')
    projeto = data.get('projeto') or request.headers.get('X-Projeto')
    colabs  = data.get('colaboradores')
    relato  = data.get('relatorio')
    pct_raw = data.get('porcentagem')

    if numero is None or sheet is None:
        return jsonify({'success': False, 'error': 'Parâmetros insuficientes'}), 400

    if not projeto:
        projeto = _descobrir_projeto_por_tarefa(int(numero), str(sheet))
        if projeto in (None, "__AMBIGUO__"):
            return jsonify({'success': False, 'error': 'Projeto não encontrado/ambíguo'}), 409

    df = carregar_base_dados(projeto)
    if df is None or df.empty:
        return jsonify({'success': False, 'error': 'Projeto sem dados'}), 404

    m = (df['numero'].astype(str) == str(numero)) & (df['Sheet'].astype(str) == str(sheet))
    idxs = df.index[m].tolist()
    if not idxs:
        return jsonify({'success': False, 'error': 'Tarefa não encontrada'}), 404
    i = idxs[0]

    # precisa estar em curso
    if int(df.at[i, 'em_curso'] or 0) != 1:
        return jsonify({'success': False, 'error': 'not_in_progress'}), 409

    op = _current_operator()
    dono = str(df.at[i, 'em_curso_by'] or '').strip()
    if dono and dono.lower() != op.lower():
        return jsonify({'success': False, 'error': 'locked_by_other', 'em_curso_by': dono}), 409

    if isinstance(colabs, list):
        colabs = ','.join([str(x).strip() for x in colabs if str(x).strip()])
    if colabs is not None:
        df.at[i, 'colaboradores'] = str(colabs).strip()
    if relato is not None:
        df.at[i, 'relatorio_progresso'] = str(relato).strip()

    if pct_raw is not None:
        try:
            p = int(str(pct_raw).replace('%','').strip() or 0)
        except Exception:
            p = 0
        df.at[i, 'porcentagem'] = max(0, min(99, p))  # 100% = concluir pelo fluxo correto

    # salva
    try:
        projetos = _listar_projetos()
        target   = projetos[projeto]
        with pd.ExcelWriter(target, engine='openpyxl') as w:
            df_proj = df[df['Projeto'].astype(str) == str(projeto)]
            for aba, gdf in df_proj.groupby('Sheet'):
                cols = [c for c in gdf.columns if c not in ['Sheet','Projeto']]
                gdf[cols].to_excel(w, sheet_name=str(aba), index=False)
    except Exception as e:
        app.logger.error(f"Erro ao salvar ATUALIZAR: {e}", exc_info=True)
        return jsonify({'success': False, 'error': 'Erro ao salvar'}), 500

    try:
        audit_log({
            'action': 'atualizar_progresso',
            'operator': op,
            'sheet': sheet,
            'task_uuid': df.at[i, 'task_uuid'] if 'task_uuid' in df.columns else None,
            'after': df.loc[i].to_dict()
        })
    except Exception:
        pass

    return jsonify({'success': True, 'task': _json_safe(df.loc[i].to_dict()), 'projeto': projeto})



# -----------------------------------------------------------------------------
# DASHBOARD
# -----------------------------------------------------------------------------
@app.route('/dashboard-metrics', methods=['POST'])
def dashboard_metrics():
    data = request.get_json(force=True, silent=True) or {}
    projeto   = data.get('projeto')
    categoria = data.get('categoria', '')
    condicoes = data.get('condicoes', {})

    df = carregar_base_dados(projeto)

    if df is None or df.empty:
        return jsonify({'total': 0, 'percent_concluidas': 0, 'criticas_abertas': 0, 'atrasadas': 0,
                        'por_prioridade': {}, 'por_fase': {}, 'por_categoria': {}})

    # Normalizações seguras
    df['duracao_str'] = df.get('duracao', '').astype(str)
    df['duracao_num'] = pd.to_numeric(df.get('duracao', 0), errors='coerce').fillna(0).astype(int)
    df['pct_num']     = pd.to_numeric(df.get('porcentagem', 0), errors='coerce').fillna(0).astype(int)
    df['conc_num']    = pd.to_numeric(df.get('concluida', 0), errors='coerce').fillna(0).astype(int)

    # Sinal unificado de conclusão (qualquer um vale)
    df['is_concluida'] = (
        (df['conc_num'] >= 1) |
        (df['pct_num'] >= 100) |
        (df['duracao_str'].str.contains('conclu', case=False, na=False))
    ).astype(int)

    # Filtro por categoria (se vier)
    if categoria:
        df = df[df['categoria'] == categoria]

    # Filtro por condições (fase/condição) – igual ao seu, só reaplicado antes do cômputo
    if condicoes:
        def tarefa_valida(row):
            fase = str(row.get('fase',''))
            cond = str(row.get('condicao','')).strip()
            return fase in condicoes and cond in (condicoes.get(fase) or [])
        df = df[df.apply(tarefa_valida, axis=1)]

    total       = int(len(df))
    concluidas  = int(df['is_concluida'].sum())
    criticas_ab = int(len(df[(df['is_concluida'] != 1) & (df['condicao'] == 'Sempre')]))
    atrasadas   = int(len(df[(df['is_concluida'] != 1) & (df['duracao_num'] < 7)]))

    return jsonify({
        'total': total,
        'percent_concluidas': round((concluidas/total*100), 1) if total else 0,
        'criticas_abertas': criticas_ab,
        'atrasadas': atrasadas,
        'por_prioridade': df.groupby('condicao').size().to_dict(),
        'por_fase': df.groupby('fase').size().to_dict(),
        'por_categoria': df.groupby('categoria').size().to_dict()
    })

# -----------------------------------------------------------------------------
# ADICIONAR TAREFA
# -----------------------------------------------------------------------------
@app.route('/adicionar-tarefa', methods=['POST'])
def adicionar_tarefa():
    """
    Cria uma nova tarefa no projeto/aba informados.
    Mantém o fluxo essencial:
      - normalização de payload
      - numeração por aba
      - escrita via pandas por projeto/aba
    Adições:
      - task_uuid (v4) e version=1
      - project_uuid (determinístico por projeto)
      - prevenção de numero duplicado (409 com sugestão)
      - validação tolerante de esquema (400 com details)
      - backup rotativo do XLSX do projeto
      - auditoria (before=None -> after=linha criada)
    """
    try:
        data = request.get_json(force=True) if request.is_json else (request.get_json() or {})

        # ---------- Normalizações de chaves (compatibilidade com front) ----------
        # 'Classificação' -> 'classificacao'
        if 'classificacao' not in data and 'Classificação' in data:
            data['classificacao'] = data.pop('Classificação')
        # 'Condição' -> 'condicao'
        if 'condicao' not in data and 'Condição' in data:
            data['condicao'] = data.pop('Condição')

        # ---------- Campos básicos ----------
        projeto = (data.get('projeto') or '').strip()
        if not projeto:
            return jsonify({'success': False, 'error': 'Informe o projeto'}), 400

        aba = (data.get('sheet') or data.get('aba') or 'Backlog').strip() or 'Backlog'

        # Nome é obrigatório
        nome = str(data.get('nome') or data.get('Nome') or '').strip()
        if not nome:
            return jsonify({'success': False, 'error': 'Informe o nome da tarefa'}), 400

        # Fase com default seguro (mantém seu fluxo; pode vir livre)
        fase = str(data.get('fase') or data.get('Fase') or 'Aberta').strip()

        # O resto é opcional (continuam aceitando como antes)
        condicao      = str(data.get('condicao') or '').strip()             # opcional (validada só se vier)
        classificacao = str(data.get('classificacao') or '').strip()        # opcional (não é prioridade)
        categoria     = str(data.get('categoria') or '').strip()
        como_fazer    = str(data.get('como_fazer') or '').strip()
        doc_ref       = str(data.get('documento_referencia') or '').strip()

        # Duração e % (mantém lógica atual, com tolerância)
        dur = data.get('duracao', 0)
        if isinstance(dur, str) and dur.isdigit():
            dur = int(dur)
        elif isinstance(dur, (int, float)):
            pass
        else:
            # se vier texto tipo "1", vira 1; se vier vazio, vira 0
            try:
                dur = int(str(dur).strip() or 0)
            except Exception:
                dur = 0

        pct_raw = str(data.get('porcentagem', 0))
        try:
            pct = int(pct_raw.replace('%', '').strip() or 0)
        except Exception:
            pct = 0

        # ---------- Verificações de projeto e carga do DF ----------
        projetos = _listar_projetos()
        if projeto not in projetos:
            return jsonify({'success': False, 'error': 'Projeto inválido'}), 400

        # carrega DF do projeto (já garante colunas extras e uuids na carga)
        df = carregar_base_dados(projeto)

        # ---------- Numeração (previne duplicata e sugere número livre) ----------
        existentes = set(
            pd.to_numeric(
                df.loc[(df['Projeto'] == projeto) & (df['Sheet'] == aba), 'numero'],
                errors='coerce'
            ).dropna().astype(int).tolist()
        )

        if data.get('numero') is not None:
            try:
                numero = int(str(data.get('numero')))
            except Exception:
                return jsonify({'success': False, 'error': 'Número inválido'}), 400
            if numero in existentes:
                return jsonify({
                    'success': False,
                    'error': 'numero_duplicado',
                    'message': f'Número {numero} já existe em {projeto}/{aba}.',
                    'suggested_numero': (max(existentes) + 1) if existentes else 1
                }), 409
        else:
            # escolhe o próximo LIVRE (não apenas max+1) — evita colisões em buracos
            numero = 1
            while numero in existentes:
                numero += 1

        # ---------- Montagem da nova linha (mantendo seu esquema) ----------
        nova = {
            'numero': numero,
            'classificacao': classificacao,         # opcional; NÃO é prioridade
            'categoria': categoria,
            'fase': fase,                           # pode ser livre
            'condicao': condicao,                   # opcional; se vier, validada (Sempre/A/B/C)
            'nome': nome,
            'duracao': dur,
            'como_fazer': como_fazer,
            'documento_referencia': doc_ref,
            'porcentagem': pct,
            'concluida': 0,
            'Sheet': aba,
            'Projeto': projeto,

            # >>> NOVO: segurança/consistência por tarefa <<<
            'task_uuid': str(uuid.uuid4()),
            'version': 1,
            'project_uuid': _project_uuid(projeto),
        }

        # Log útil para debug (não polui produção)
        app.logger.debug(f"ADD nova linha: {{'numero': {numero}, 'nome': '{nome}', 'fase': '{fase}', "
                         f"'condicao': '{condicao}', 'classificacao': '{classificacao}', 'duracao': {dur}, "
                         f"'Sheet': '{aba}', 'Projeto': '{projeto}'}}")

        # ---------- Validação tolerante ----------
        try:
            validate_row(nova)
        except ValidationError as ve:
            app.logger.warning(f"Validação falhou ao adicionar: {ve.errors}")
            return jsonify({'success': False, 'error': 'validation', 'details': ve.errors}), 400

        # ---------- Append no DF do projeto (mantendo sua abordagem com pandas) ----------
        df = pd.concat([df, pd.DataFrame([nova])], ignore_index=True)

        # ---------- Persistência com backup rotativo (somente o XLSX do projeto) ----------
        target = projetos[projeto]
        try:
            rotate_backups(target)
        except Exception:
            # backup falhou não deve impedir a escrita — apenas registra log
            app.logger.exception("Falha ao criar backup rotativo antes de salvar")

        with pd.ExcelWriter(target, engine='openpyxl') as w:
            for sname, gdf in df[df['Projeto'] == projeto].groupby('Sheet'):
                cols = [c for c in gdf.columns if c not in ['Sheet', 'Projeto']]
                gdf[cols].to_excel(w, sheet_name=sname, index=False)

        # ---------- Auditoria ----------
        audit_log({
            'action': 'adicionar',
            'operator': data.get('operator') or request.headers.get('X-Operator') or 'system',
            'sheet': aba,
            'task_uuid': nova['task_uuid'],
            'before': None,
            'after': nova
        })

        return jsonify({'success': True, 'sheet': aba, 'projeto': projeto, 'task': nova}), 200

    except ValidationError as ve:
        # fallback (se por alguma razão levantar aqui)
        return jsonify({'success': False, 'error': 'validation', 'details': ve.errors}), 400

    except Exception as e:
        app.logger.exception("Erro ao adicionar nova tarefa")
        return jsonify({'success': False, 'error': str(e)}), 500

# -----------------------------------------------------------------------------
# LISTAR ABAS & CATEGORIAS
# -----------------------------------------------------------------------------
@app.route('/listar-abas')
def listar_abas():
    projeto = request.args.get('projeto') or request.headers.get('X-Projeto')
    projetos = _listar_projetos()
    if projeto and projeto in projetos:
        planilha = pd.read_excel(projetos[projeto], sheet_name=None, engine="openpyxl")
        return jsonify(list(planilha.keys()))
    # sem projeto → retorna dict {projeto:[abas]}
    resp = {}
    for nome, path in projetos.items():
        try:
            sheets = pd.read_excel(path, sheet_name=None, engine="openpyxl").keys()
            resp[nome] = list(sheets)
        except:
            resp[nome] = []
    return jsonify(resp)

@app.route('/categorias-usadas')
def categorias_usadas():
    try:
        projeto = request.args.get('projeto') or request.headers.get('X-Projeto')
        df = carregar_base_dados(projeto)
        cats = sorted([str(c).strip() for c in df['categoria'].dropna().unique().tolist() if str(c).strip()])
        return jsonify(cats)
    except Exception as e:
        app.logger.error(f"categorias-usadas: {e}", exc_info=True)
        return jsonify([]), 500

def _descobrir_projeto_por_tarefa(numero: int, sheet: str | None):
    """Procura em TODOS os .xlsx e retorna o nome do projeto em que (numero, sheet) existe.
       Se encontrar único match → retorna str. Se nenhum → None. Se mais de um → "__AMBIGUO__".
    """
    df_all = carregar_base_dados(projeto=None)
    if df_all is None or df_all.empty:
        return None
    m = (df_all['numero'] == int(numero))
    if sheet:
        m = m & (df_all['Sheet'].astype(str) == str(sheet))
    hits = df_all.loc[m, 'Projeto'].dropna().unique().tolist()
    if not hits:
        return None
    if len(hits) == 1:
        return hits[0]
    return "__AMBIGUO__"


# -----------------------------------------------------------------------------
# PROJETOS
# -----------------------------------------------------------------------------
@app.route('/exportar-projeto')
def exportar_projeto():
    from io import BytesIO
    import pandas as pd
    from flask import send_file

    projeto = request.args.get('projeto')
    formato = request.args.get('formato', 'xlsx').lower()

    if not projeto:
        return "Projeto não especificado", 400

    df = carregar_base_dados(projeto)
    if df is None or df.empty:
        return "Projeto não encontrado ou sem dados", 404

    output = BytesIO()

    if formato == 'xlsx':
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name=projeto[:31])
        mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        ext = 'xlsx'
    elif formato == 'pdf':
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet
        styles = getSampleStyleSheet()
        doc = SimpleDocTemplate(output, pagesize=A4)
        data = [df.columns.tolist()] + df.values.tolist()
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#d63384')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.gray),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold')
        ]))
        elements = [Paragraph(f'Projeto: {projeto}', styles['Heading2']), table]
        doc.build(elements)
        mimetype = 'application/pdf'
        ext = 'pdf'
    else:
        return "Formato inválido", 400

    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=f"{projeto}.{ext}",
        mimetype=mimetype
    )


@app.route('/listar-projetos')
def listar_projetos():
    return jsonify(sorted(_listar_projetos().keys()))

def _listar_projetos():
    """Retorna {nome_projeto: caminho_arquivo} para todos .xlsx.
    1ª preferência: /data; se vazio, cai para a raiz do projeto.
    """
    projetos = {}
    # 1) primeiro tenta /data
    paths = glob.glob(os.path.join(DATA_DIR, "*.xlsx"))
    # 2) se vazio, tenta a raiz do projeto (onde costuma ficar o Conjuntas.xlsx)
    if not paths:
        paths = glob.glob(os.path.join(BASE_DIR, "*.xlsx"))

    for path in paths:
        nome = os.path.splitext(os.path.basename(path))[0]
        projetos[nome] = path
    return projetos

from openpyxl import Workbook
from werkzeug.utils import secure_filename

COLS_PADRAO = ['Número','Classificação','Categoria','Fase','Condição','Nome','Duração','Como Fazer','Documento Referência','% Concluída']


@app.route('/criar-projeto', methods=['POST'])
def criar_projeto():
    data = request.get_json(force=True, silent=True) or {}
    nome = (data.get('nome') or '').strip()
    if not nome:
        return jsonify({'success': False, 'error': 'Informe um nome'}), 400

    # garante .xlsx e caminho seguro
    if not nome.lower().endswith('.xlsx'):
        nome += '.xlsx'
    fname = secure_filename(nome)
    path = os.path.join(DATA_DIR, fname)

    if os.path.exists(path):
        return jsonify({'success': False, 'error': 'Já existe um projeto com esse nome'}), 400

    # cria workbook com uma aba padrão
    wb = Workbook()
    ws = wb.active
    ws.title = 'Backlog'
    ws.append(COLS_PADRAO)
    wb.save(path)

    return jsonify({'success': True, 'projeto': os.path.splitext(fname)[0]})

@app.route('/upload-projeto', methods=['POST'])
def upload_projeto():
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'Nenhum arquivo enviado'}), 400
    f = request.files['file']
    if not f or f.filename == '':
        return jsonify({'success': False, 'error': 'Arquivo inválido'}), 400
    if not f.filename.lower().endswith('.xlsx'):
        return jsonify({'success': False, 'error': 'Apenas .xlsx é aceito'}), 400

    fname = secure_filename(f.filename)
    dest = os.path.join(DATA_DIR, fname)
    f.save(dest)

    # valida leitura (opcional): tenta abrir para garantir que não corrompeu
    try:
        _ = pd.read_excel(dest, sheet_name=None, engine='openpyxl')
    except Exception as e:
        os.remove(dest)
        return jsonify({'success': False, 'error': f'Planilha inválida: {e}'}), 400

    return jsonify({'success': True, 'projeto': os.path.splitext(fname)[0]})



# =============================================================================
# ============================== LOGIN  =====================================
# =============================================================================

def _seed_users_if_missing():
    if not os.path.exists(USERS_FILE):
        seed = [
            {"id":"func1","name":"Funcionário Demo","email":"func@demo","password":generate_password_hash("123456"),"role":"funcionario","active":True},
            {"id":"gest1","name":"Gestor Demo","email":"gestor@demo","password":generate_password_hash("123456"),"role":"gestor","active":True},
            {"id":"sup1","name":"Supervisor Demo","email":"supervisor@demo","password":generate_password_hash("123456"),"role":"supervisor","active":True},
        ]
        with open(USERS_FILE, "w", encoding="utf-8") as f:
            json.dump(seed, f, ensure_ascii=False, indent=2)

def _load_users():
    _seed_users_if_missing()
    with open(USERS_FILE, "r", encoding="utf-8") as f:
        return json.load(f)


@app.post("/login")
def do_login():
    data = request.get_json(silent=True) or {}
    email = (data.get("email") or "").strip().lower()
    password = data.get("password") or ""
    users = _load_users()
    user = next((u for u in users if u["email"].lower()==email and u.get("active", True)), None)
    if not user or not check_password_hash(user["password"], password):
        return jsonify({"error":"Credenciais inválidas"}), 401
    session["user"] = {k: user[k] for k in ("id","name","email","role")}
    resp = jsonify({"ok": True, "user": session["user"]})
    # anti-cache
    resp.headers["Cache-Control"] = "no-store"
    return resp

@app.post("/logout")
def do_logout():
    session.clear()
    resp = jsonify({"ok": True})
    # remove cookie de sessão
    resp.delete_cookie("session", path="/")
    resp.headers["Cache-Control"] = "no-store"
    return resp

@app.get("/me")
def me():
    u = session.get("user")
    if not u:
        r = jsonify({"error":"unauthorized"}); r.status_code = 401
        r.headers["Cache-Control"] = "no-store"
        return r
    r = jsonify({"user": u})
    r.headers["Cache-Control"] = "no-store"
    return r

def _current_operator():
    """
    Retorna o email do operador atual.
    Prioridade: header X-Operator > sessão.user.email > sessão.user.name
    Retorna string vazia se nenhum operador válido for encontrado.
    """
    u = session.get("user") or {}
    
    # Tenta pegar do header primeiro
    op_header = request.headers.get('X-Operator', '').strip()
    
    # Valida se o header tem valor real
    if op_header and op_header not in ('', 'system', 'undefined', 'null'):
        return op_header
    
    # Fallback: sessão
    op_session = u.get('email', '').strip() or u.get('name', '').strip()
    
    # Retorna vazio se não houver operador válido (não use 'system' como fallback)
    return op_session if op_session else ''


def current_user():
    return session.get("user")  # dict: id, name, email, role

def login_required(fn):
    @wraps(fn)
    def _wrap(*args, **kwargs):
        if not current_user():
            return jsonify({"error":"unauthorized"}), 401
        return fn(*args, **kwargs)
    return _wrap

@app.post("/login")
def login():
    data = request.get_json(silent=True) or {}
    email = (data.get("email") or "").strip().lower()
    password = data.get("password") or ""
    users = _load_users()
    user = next((u for u in users if u["email"].lower()==email and u.get("active", True)), None)
    if not user or not check_password_hash(user["password"], password):
        return jsonify({"error":"Credenciais inválidas"}), 401
    session["user"] = {k: user[k] for k in ("id","name","email","role")}
    return jsonify({"ok": True, "user": session["user"]})

@app.post("/logout")
def logout():
    # Limpa a sessão
    session.clear()
    # Resposta + remoção explícita do cookie de sessão
    resp = jsonify({"ok": True})
    # nome padrão do cookie de sessão do Flask:
    resp.delete_cookie(app.session_cookie_name, path="/")
    # headers anticache
    resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    resp.headers["Pragma"] = "no-cache"
    resp.headers["Expires"] = "0"
    return resp

# ---------------------------------------------
# --------- ATENÇÃO ---------------------------
# ---------------------------------------------
@app.get("/login")
def login_page():
    return render_template("login.html")


# Opcional (global): garante no-store nas páginas principais protegidas
@app.after_request
def _no_cache(resp):
    if request.path in {"/", "/index.html", "/supervisor", "/me", "/login"}:
        resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        resp.headers["Pragma"] = "no-cache"
        resp.headers["Expires"] = "0"
    return resp












# =============================================================================
# ============================== CHATBOT  =====================================
# =============================================================================

# Stopwords para busca textual
STOP_WORDS = {
    "a","o","as","os","um","uma","uns","umas","de","do","da","dos","das","em","no","na","nos","nas",
    "por","para","com","sem","sob","sobre","que","quem","qual","onde","quando","porque","se","entao",
    "mas","ou","e","ser","estar","ter","haver","fazer","ir","vir","como","me","te","lhe","nos","vos",
    "lhes","meu","minha","teu","tua","seu","sua","nosso","vosso","deste","desta","isto","isso","aquilo",
    "pelo","pela","pelos","pelas","instrucoes","instrucao","passo"
}

def deaccent(s):
    if s is None: return ""
    s = unicodedata.normalize("NFD", str(s))
    return "".join(ch for ch in s if unicodedata.category(ch) != "Mn")

def tokenize(text):
    if pd.isna(text): return set()
    base = deaccent(str(text)).lower()
    toks = set(re.findall(r"[a-z0-9]+", base))
    return toks - STOP_WORDS

def jaccard(a, b):
    if not a and not b: return 0.0
    inter = len(a & b)
    union = len(a | b) if (a | b) else 1
    return inter / union

def score_row(query_tokens, row, fields):
    score = 0.0
    query_phrase = " ".join(sorted(query_tokens)) if query_tokens else ""
    for f in fields:
        val_norm = deaccent(str(row.get(f, ""))).lower()
        toks = tokenize(val_norm)
        score += jaccard(query_tokens, toks)
        if val_norm and query_phrase and query_phrase in val_norm:
            score += 0.1
    return score

def status_label(p01):
    if pd.isna(p01) or p01 <= 0: return "Não iniciada"
    if p01 >= 1.0: return "Finalizada"
    return "Em andamento"

def _detect_sep_from_text(text: str) -> str:
    sample = text.splitlines()[:200]
    comma = sum(l.count(",") for l in sample if l.strip())
    semi  = sum(l.count(";") for l in sample if l.strip())
    return "," if comma >= semi else ";"

def _fix_broken_lines(text: str, quotechar: str = '"') -> str:
    out, buf, open_quotes = [], [], 0
    for line in text.splitlines():
        buf.append(line)
        open_quotes += line.count(quotechar)
        if open_quotes % 2 == 0:
            out.append(" ".join(buf))
            buf, open_quotes = [], 0
    if buf: out.append(" ".join(buf))
    return "\n".join(out)

def safe_read_csv(path: str):
    encodings = ["utf-8-sig","utf-8","latin-1"]
    last_err = None
    for enc in encodings:
        try:
            with open(path,"r",encoding=enc,errors="replace") as f:
                raw = f.read()
            fixed = _fix_broken_lines(raw)
            sep = _detect_sep_from_text(fixed)
            df = pd.read_csv(StringIO(fixed), sep=sep, engine="python",
                             quotechar='"', doublequote=True, on_bad_lines="skip")
            return df, sep, enc
        except Exception as e:
            last_err = e
    raise RuntimeError(f"Falha ao ler CSV: {last_err}")

def repair_malformed_rows(df: pd.DataFrame, sep: str, expected_cols):
    if "Projeto" not in df.columns: return 0
    for c in expected_cols:
        if c not in df.columns: df[c] = ""
    others = [c for c in expected_cols if c != "Projeto"]

    def row_is_bad(idx):
        proj = str(df.at[idx, "Projeto"])
        if sep in proj: return False
        for c in others:
            val = df.at[idx, c]
            if pd.notna(val) and str(val).strip() != "": return False
        return True

    bad_idxs = [idx for idx in df.index if row_is_bad(idx)]
    fixed = 0
    for idx in bad_idxs:
        raw = str(df.at[idx, "Projeto"])
        try:
            parsed = next(csv.reader([raw], delimiter=sep, quotechar='"', doublequote=True))
            if len(parsed) < len(expected_cols): parsed += [""] * (len(expected_cols) - len(parsed))
            else: parsed = parsed[:len(expected_cols)]
            for col, val in zip(expected_cols, parsed):
                df.at[idx, col] = val
            fixed += 1
        except Exception:
            pass
    return fixed

DF = None
SEARCH_FIELDS = ["Nome","ComoFazer","Categoria","Fase","Condicao"]

def load_dataset():
    """Carrega 'Conjuntas_combinado.csv' da raiz do app."""
    global DF
    candidates = [
        os.path.join(BASE_DIR, "Conjuntas_combinado.csv"),
        "Conjuntas_combinado.csv"
    ]
    csv_path = next((p for p in candidates if os.path.exists(p)), None)
    if not csv_path:
        print("ERRO: Coloque 'Conjuntas_combinado.csv' na pasta do app.")
        DF = None
        return

    df, sep, enc = safe_read_csv(csv_path)
    expected = ["Projeto","Numero","Classificacao","Categoria","Fase","Condicao",
                "Nome","Duracao","ComoFazer","DocumentoReferencia","PctConcluida"]

    # renomeia colunas por nome "normalizado"
    norm_map = {c.lower().strip(): c for c in df.columns}
    for col in expected:
        key = col.lower()
        if col not in df.columns and key in norm_map:
            df = df.rename(columns={norm_map[key]: col})
    for col in expected:
        if col not in df.columns: df[col] = ""

    repair_malformed_rows(df, sep, expected)

    df["Projeto"] = df["Projeto"].fillna("").replace("", "Sem Projeto")

    dur_raw = df["Duracao"].astype(str).str.replace(",", ".", regex=False)
    dur_raw = dur_raw.str.extract(r"([-+]?\d*\.?\d+)")[0]
    df["_dur"] = pd.to_numeric(dur_raw, errors="coerce")

    pct_raw = df["PctConcluida"].astype(str).str.replace("%","",regex=False).str.replace(",",".",regex=False)
    pct_raw = pct_raw.str.extract(r"([-+]?\d*\.?\d+)")[0]
    pct = pd.to_numeric(pct_raw, errors="coerce")
    df["_pct01"]  = pct.where(pct <= 1.0, pct / 100.0)
    df["_pct100"] = df["_pct01"] * 100.0

    df[SEARCH_FIELDS] = df[SEARCH_FIELDS].fillna("")
    DF = df
    print("Base 'Conjuntas_combinado.csv' carregada.")

# carrega dataset na inicialização
try:
    load_dataset()
except Exception as e:
    print("Falha ao carregar dataset:", e)
    DF = None

def reply_search(query: str):
    if DF is None:
        return (False, "Base de dados não disponível.")
    qtokens = tokenize(query)
    scores = [(i, score_row(qtokens, row, SEARCH_FIELDS)) for i, row in DF.iterrows()]
    scores.sort(key=lambda x: x[1], reverse=True)
    SCORE_MIN = 0.35
    if not scores or scores[0][1] < SCORE_MIN:
        return (False, "🤖 Não encontrei nada parecido com sua busca. Tente outras palavras ou digite 'ajuda'.")
    i, _ = scores[0]
    row = DF.iloc[i]
    p01 = row.get("_pct01")
    dur = row.get("_dur")
    resposta = (f"🤖 Instruções da tarefa\n"
                f"Tarefa: {row.get('Nome')}\n"
                f"Etapa: {row.get('Fase')}\n"
                f"Status: {status_label(p01)}  |  Andamento: {('-' if pd.isna(p01) else f'{p01*100:.1f}%')}  "
                f"|  Duração: {('-' if pd.isna(dur) else f'{dur:.2f}')}\n\n"
                f"Passo a passo:\n{row.get('ComoFazer')}\n\n"
                f"📎 Documento: {row.get('DocumentoReferencia')}")
    return (True, resposta)

def reply_status_project() -> str:
    if DF is None: return "Base de dados não disponível."
    g = DF.groupby("Projeto", dropna=False)
    total_qtde = g.size()
    tempo_total = g["_dur"].sum(min_count=1)
    media_and = g["_pct01"].mean() * 100.0
    concl = g["_pct01"].apply(lambda s: (s.fillna(0) >= 1.0).sum())
    em_and = g["_pct01"].apply(lambda s: ((s.fillna(0) > 0) & (s.fillna(0) < 1.0)).sum())
    nao_ini = g["_pct01"].apply(lambda s: (s.fillna(0) == 0).sum())
    media_nao = (1 - (concl / total_qtde.replace(0, pd.NA))) * 100.0
    res = pd.DataFrame({
        "Tarefas": total_qtde, "Tempo total": tempo_total, "Andamento médio (%)": media_and,
        "Não finalizado médio (%)": media_nao, "Finalizadas": concl, "Em andamento": em_and, "Não iniciadas": nao_ini
    }).sort_values("Não finalizado médio (%)", ascending=False)
    lines = ["📊 Andamento por projeto"]
    for n, (proj, row) in enumerate(res.iterrows(), start=1):
        ttot = "-" if pd.isna(row["Tempo total"]) else f"{row['Tempo total']:.2f}"
        am   = "-" if pd.isna(row["Andamento médio (%)"]) else f"{row['Andamento médio (%)']:.1f}%"
        nf   = "-" if pd.isna(row["Não finalizado médio (%)"]) else f"{row['Não finalizado médio (%)']:.1f}%"
        lines.append(f"{n}) {proj} — Tarefas: {int(row['Tarefas'])} | Tempo: {ttot} | "
                     f"Andamento médio: {am} | Não finalizado médio: {nf} | "
                     f"Finalizadas: {int(row['Finalizadas'])} | Em andamento: {int(row['Em andamento'])} | "
                     f"Não iniciadas: {int(row['Não iniciadas'])}")
    return "\n".join(lines)

def reply_status_stage() -> str:
    if DF is None: return "Base de dados não disponível."
    g = DF.groupby("Fase", dropna=False)
    total = g.size()
    concl = g["_pct01"].apply(lambda s: (s.fillna(0) >= 1.0).sum())
    em_and = g["_pct01"].apply(lambda s: ((s.fillna(0) > 0) & (s.fillna(0) < 1.0)).sum())
    nao_ini = g["_pct01"].apply(lambda s: (s.fillna(0) == 0).sum())
    media_pct = g["_pct100"].mean()
    media_nao = (1 - (concl / total.replace(0, pd.NA))) * 100.0
    tempo_total = g["_dur"].sum(min_count=1)
    tempo_conc = (DF["_dur"].fillna(0) * DF["_pct01"].fillna(0)).groupby(DF["Fase"]).sum(min_count=1)
    tempo_pend = (DF["_dur"].fillna(0) * (1 - DF["_pct01"].fillna(0))).groupby(DF["Fase"]).sum(min_count=1)
    res = pd.DataFrame({
        "Tarefas": total, "Finalizadas": concl, "Em andamento": em_and, "Não iniciadas": nao_ini,
        "Andamento médio (%)": media_pct, "Não finalizado médio (%)": media_nao,
        "Tempo total": tempo_total, "Tempo concluído": tempo_conc, "Tempo a concluir": tempo_pend
    }).sort_values(["Em andamento","Não iniciadas"], ascending=False)
    lines = ["📉 Andamento por etapa"]
    for n, (fase, row) in enumerate(res.iterrows(), start=1):
        at = "-" if pd.isna(row["Andamento médio (%)"]) else f"{row['Andamento médio (%)']:.1f}%"
        nf = "-" if pd.isna(row["Não finalizado médio (%)"]) else f"{row['Não finalizado médio (%)']:.1f}%"
        tt = "-" if pd.isna(row["Tempo total"]) else f"{row['Tempo total']:.2f}"
        tc = "-" if pd.isna(row["Tempo concluído"]) else f"{row['Tempo concluído']:.2f}"
        tp = "-" if pd.isna(row["Tempo a concluir"]) else f"{row['Tempo a concluir']:.2f}"
        lines.append(f"{n}) {fase} — Tarefas: {int(row['Tarefas'])} | Finalizadas: {int(row['Finalizadas'])} | "
                     f"Em andamento: {int(row['Em andamento'])} | Não iniciadas: {int(row['Não iniciadas'])} | "
                     f"Andamento médio: {at} | Não finalizado médio: {nf} | Tempo total: {tt} | "
                     f"Concluído: {tc} | A concluir: {tp}")
    return "\n".join(lines)

def reply_pending(limit=5):
    if DF is None: return "Base de dados não disponível."
    dfp = DF.copy()
    dfp["_faltante"] = dfp["_dur"].fillna(0) * (1 - dfp["_pct01"].fillna(0))
    dfp = dfp[dfp["_faltante"] > 0]
    if dfp.empty: return "✅ Não há tarefas em aberto."
    dfp = dfp.sort_values("_faltante", ascending=False)
    if limit and limit > 0:
        top = dfp.head(limit); data = top; title = f"📝 Tarefas pendentes (Top {len(top)}) — ordenadas por tempo que ainda falta"
    else:
        data = dfp; title = "📝 Tarefas pendentes — TODAS (ordenadas por tempo que ainda falta)"
    lines = [title]
    for _, r in data.iterrows():
        andamento = "-" if pd.isna(r["_pct100"]) else f"{r['_pct100']:.1f}%"
        tempo = "-" if pd.isna(r["_dur"]) else f"{r['_dur']:.2f}"
        falta = f"{r['_faltante']:.2f}"
        lines.append(f"• {r['Nome']} — Projeto: {r['Projeto']} | Etapa: {r['Fase']} | "
                     f"Andamento: {andamento} | Duração: {tempo} | Falta: {falta}")
    return "\n".join(lines)

def reply_filter_by_condition(condition: str) -> str:
    if DF is None: return "Base de dados não disponível."
    if 'Condicao' not in DF.columns: return "ERRO: A coluna 'Condicao' não foi encontrada no CSV."
    filtered_df = DF[DF['Condicao'].str.lower() == condition.lower()]
    if filtered_df.empty: return f"😕 Nenhuma tarefa encontrada com a condição '{condition}'."
    lines = [f"🔎 Tarefas com a condição '{condition.title()}' ({len(filtered_df)} encontradas):"]
    for _, row in filtered_df.iterrows():
        lines.append(f"• {row.get('Nome','(sem nome)')} (Projeto: {row.get('Projeto','-')}, Etapa: {row.get('Fase','-')})")
    return "\n".join(lines)

def reply_task_summary_by_conditions() -> dict:
    if DF is None: return {"error": "Base de dados não disponível."}
    if 'Condicao' not in DF.columns or 'Nome' not in DF.columns:
        return {"error": "ERRO: Colunas 'Condicao' ou 'Nome' ausentes no CSV."}
    conditions = ['Sempre','A','B','C']
    summary = {}
    for c in conditions:
        summary[c] = DF[DF['Condicao'].str.lower() == c.lower()]['Nome'].tolist()
    return summary

def reply_all_tasks() -> str:
    if DF is None or DF.empty: return "A base de dados está vazia ou não foi carregada."
    lines = [f"📋 Lista de Todas as Tarefas ({len(DF)} no total):"]
    for projeto, grupo in DF.groupby('Projeto'):
        lines.append(f"\n--- Projeto: {projeto} ---")
        for _, t in grupo.iterrows():
            lines.append(f"• {t.get('Nome','(sem nome)')} (Etapa: {t.get('Fase','-')})")
    return "\n".join(lines)

def extract_condition_from_query(query: str):
    valid = ['A','B','C','Sempre']
    pattern = re.compile(r"\b(" + "|".join(valid) + r")\b", re.IGNORECASE)
    match = pattern.search(query or "")
    return match.group(1) if match else None

def bot_reply(message: str):
    q = (message or "").strip().lower()
    if not q:
        return "Olá! Digite 'ajuda' para ver todas as opções."
    if q in {"ajuda","help","?"}:
        return ("Posso ajudar com:\n"
                "- **Mostrar Todas as Tarefas**: digite `mostrar todas as tarefas`\n"
                "- **Resumo de Tarefas**: digite `filtrar tarefas`\n"
                "- **Filtrar por Condição**: ex: `filtrar por A` ou `tarefas com status C`\n"
                "- **Status por Projeto**: digite `status por projeto`\n"
                "- **Status por Etapa**: digite `status por etapa`\n"
                "- **Tarefas Pendentes (Top 5)**: digite `tarefas pendentes`\n"
                "- **Buscar Instruções**: digite o nome da tarefa")
    if "todas" in q and "tarefas" in q:
        return reply_all_tasks()
    if q == "filtrar tarefas":
        return reply_task_summary_by_conditions()
    if "status" in q and ("projeto" in q or "projetos" in q):
        return reply_status_project()
    if "status" in q and ("etapa" in q or "fase" in q):
        return reply_status_stage()
    if ("tarefas" in q and ("aberto" in q or "pendente" in q or "pendentes" in q)):
        return reply_pending(limit=0) if "todas" in q else reply_pending(limit=5)
    cond = extract_condition_from_query(q)
    if cond:
        return reply_filter_by_condition(cond)
    ok, resp = reply_search(message)
    return resp

@app.route("/chatbot")
def chatbot_page():
    return render_template("chatbot.html")

@app.route("/api/chat", methods=["POST"])
def api_chat():
    data = request.get_json(force=True, silent=True) or {}
    message = data.get("message") or data.get("pergunta") or data.get("msg")
    if not message:
        return jsonify({"ok": False, "error": "Nenhuma mensagem recebida."}), 400
    try:
        reply = bot_reply(message)
        # Caso a resposta seja um dicionário (resumo por condição), devolvemos como tal;
        # caso contrário, string simples.
        return jsonify({"ok": True, "reply": reply})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/admin/reload", methods=["POST","GET"])
def reload_base():
    try:
        load_dataset()
        return jsonify({"ok": True, "msg": "Base recarregada com sucesso."})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

# -----------------------------------------------------------------------------
# MAIN
# -----------------------------------------------------------------------------
if __name__ == '__main__':
    app.run(debug=True)
